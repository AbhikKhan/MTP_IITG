# import networkx as nx
# import matplotlib.pyplot as plt
# from networkx.drawing.nx_agraph import graphviz_layout

# # def dtree(treelist):
# #     # Create an empty directed graph
# #     G = nx.DiGraph()

# #     # Add edges to the graph using the treelist
# #     for sublist in treelist:
# #         if len(sublist) != 3:
# #             print("Invalid sublist in treelist: " + str(sublist))
# #             continue

# #         edge_value, u, v = sublist
# #         G.add_edge(u, v, weight=edge_value)

# #     # Set positions for the tree nodes using Graphviz layout
# #     pos = graphviz_layout(G, prog='dot')

# #     # Draw the graph with bold fonts
# #     nx.draw(G, pos, with_labels=True, node_size=500, node_color='lightblue', font_size=10, font_weight='bold', edge_color='gray', width=1.5, arrows=True, arrowstyle='->', arrowsize=10)

# #     # Add edge labels
# #     labels = nx.get_edge_attributes(G, 'weight')
# #     nx.draw_networkx_edge_labels(G, pos, edge_labels=labels)

# #     # Show the graph
# #     plt.axis('off')
# #     plt.show()


# import pydot

# def _createTree(treevec, label=None):
#     '''
#     convert a tree from NODE data structure to Pydot Graph for visualisation
#     '''
#     P = pydot.Dot(graph_type='digraph', label=label, labelloc='top', labeljust='left')#, nodesep="1", ranksep="1")
    
#     nodelist, edgelist = getNodesEdges(root)
#     # pprint("NODELIST",nodelist)
#     # pp.pprint(edgelist)
    
#     var=0
    
#     # Nodes
#     for node in nodelist:
#         n = pydot.Node(node[0], label=node[1])
#         P.add_node(n)
    
    
#     for ee in range(len(treevec)):
        
#             # list=[1,treelist[1][e][0][1],"R"+str(treelist[1][e][1][1])]#[edgeValue,u,v]
#         if len(treevec[ee][1])!=3:
            
#             n = pydot.Node(var,label=treevec[ee][1])
#             P.add_node(n)
#             var=var+1
#             n = pydot.Node(var,label=treevec[ee][2])
#             P.add_node(n)
#             e = pydot.Edge(*(var-1, var), label=1, dir='back')
#             var=var+1
#         elif len(treevec[ee][1])==3 and len(treevec[ee][1]) :
#             # global nodereagents
#             nodereagents[int(treevec[ee][1][1] + treevec[ee][1][2])].append(treevec[ee][2])
#             n = pydot.Node(var,label=treevec[ee][1])
#             P.add_node(n)
#             var=var+1
#             n = pydot.Node(var,label=treevec[ee][2])
#             P.add_node(n)
#             e = pydot.Edge(*(var-1, var), label=1, dir='back')
#             var=var+1
#     # Edges
#     for edge in edgelist:
#         e = pydot.Edge(*(edge[0][0], edge[1][0]), label=edge[1][2], dir='back')
#         P.add_edge(e)
        
#     for ee in range(len(treevec)):
        
#             # list=[1,treelist[1][e][0][1],"R"+str(treelist[1][e][1][1])]#[edgeValue,u,v]
#         if len(treevec[ee][1])!=3:
#             # global nodereagents
#             e = pydot.Edge(*((treevec[ee][1][1]),(treevec[ee][2])),label=treevec[ee][1][1]), dir='back')
#         elif len(treevec[ee][1])==3:
#             # global nodereagents
#             nodereagents[int(treevec[ee][1][1] + treevec[ee][1][2])].append(treevec[ee][2])
#     return P 

# # ###############################################################################################

# # #############################################################################################

# # from IPython.display import Image, display
# # def _viewPydot(pydot):
# #     '''
# #     generates a visual plot of Pydot Graph
# #     '''
# #     pydot.write_png("output.png")
# #     pydot.write_raw("output_raw.dot")
# #     plt = Image(pydot.create_png(prog='dot'))
# #     display(plt)
    


def dtree(treevec):
    node_ids = {}
    id_counter = 0

    dot_file = "modified_graph.dot"
    r=0
    with open(dot_file,"w") as file:
        file.truncate(0)
        
    with open(dot_file, "w") as f:
        f.write("digraph G {\n")
        f.write("tree[label=< <B>MODIFIED TREE</B>> shape=RECTANGLE style=BOLD ]\n")
        for edge in treevec:
            e, u, v = edge
            if u not in node_ids:
                id_counter += 1
                node_ids[u] = id_counter
            if v not in node_ids:
                if v[0]=="R":
                    r=r+100
                
                else:
                    
                    id_counter += 1
                    node_ids[v] = id_counter
            f.write('    {} [label=< <B>{}</B>>];\n'.format(node_ids[u], u))
            if v[0]=="R":
                f.write('    {} [label=< <B>{}</B>>];\n'.format(r ,v))
                f.write('    {} -> {} [label=< <B>{}</B>>];\n'.format(node_ids[u], r , e))
            else :
                f.write('    {} [label=< <B>{}</B>>];\n'.format(node_ids[v] ,v))
                f.write('    {} -> {} [label=< <B>{}</B>>];\n'.format(node_ids[u], node_ids[v] , e))
            
            
        f.write("}")
        
        
import graphviz

import pydot

def generate_png_from_dot(dot_file_path, output_png_path):
    graph = pydot.graph_from_dot_file(dot_file_path)
    graph[0].write_png(output_png_path)


from PIL import Image

def connect_images(image1_path, image2_path, output_path):
    image1 = Image.open(image1_path)
    image2 = Image.open(image2_path)

    # Calculate the width and height of the final image
    total_width = image1.width + image2.width
    max_height = max(image1.height, image2.height)

    # Create a new blank image with the calculated dimensions
    new_image = Image.new('RGB', (total_width, max_height))

    # Paste the first image on the left side
    new_image.paste(image1, (0, 0))

    # Paste the second image on the right side
    new_image.paste(image2, (image1.width, 0))

    # Save the final image
    new_image.save(output_path)


from PIL import Image

def display_image(image_path):
    image = Image.open(image_path)
    image.show()
    
    
    
    
   
import pandas as pd

import pandas as pd
from openpyxl import load_workbook

import pandas as pd
from openpyxl import load_workbook

def save_data_to_excel(data, file_path):
    # Define the column order
    column_order = ['MIXTURE', 'MIXTURE LEN', 'TIME', 'WASTE', 'INDIVIDUAL LOADING CYCLES', 'OPTIMIZED LOADING CYCLES', 'CELLS IN FOOTPRINT']
    
    # Create a DataFrame object with specified column order
    df = pd.DataFrame(data, columns=column_order)
    
    # Load the existing Excel file
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl') 
    writer.book = book
    
    # Append the data to the existing sheet or create a new sheet
    sheet_name = 'Sheet1'  # Change to the desired sheet name
    try:
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)
    except KeyError:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Save the changes
    writer.save()
    writer.close()
